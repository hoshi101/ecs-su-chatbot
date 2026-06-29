#!/usr/bin/env python3
"""Run the 50-question API evidence set and export report files."""

from __future__ import annotations

import argparse
import csv
import json
import statistics
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

import requests


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
DEFAULT_CASES = PROJECT_ROOT / "docs" / "reports" / "department_chatbot_test_questions.csv"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "docs" / "reports"


def load_cases(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != 50:
        raise ValueError(f"Expected exactly 50 test cases, found {len(rows)} in {path}")
    return rows


def split_keywords(value: str) -> list[str]:
    return [item.strip() for item in (value or "").split(",") if item.strip()]


def trace_summary(trace_events: list[dict[str, Any]]) -> tuple[str, str]:
    if not trace_events:
        return "", ""

    route = ""
    intent = ""
    for event in trace_events:
        details = event.get("details") or {}
        if not intent:
            intent = details.get("precheck_intent") or details.get("intent") or ""
        node_name = event.get("node_name") or ""
        if node_name and node_name not in {"start", "__start__"}:
            route = node_name
    return route, intent


def run_case(
    case: dict[str, str],
    *,
    base_url: str,
    provider: str,
    model: str,
    timeout: float,
    transport: str,
    test_client: Any = None,
) -> dict[str, Any]:
    payload = {
        "query": case["query"],
        "session_id": f"api-50-evidence-{case['id']}-{uuid.uuid4()}",
        "conversation_history": [],
        "enable_web_search": True,
        "force_web_search": False,
        "similarity_threshold": 0.7,
        "llm_provider": provider,
        "llm_model": model,
    }

    started = time.perf_counter()
    status_code = 0
    error = ""
    data: dict[str, Any] = {}
    try:
        if transport == "testclient":
            response = test_client.post("/chat/", json=payload)
        else:
            response = requests.post(f"{base_url.rstrip('/')}/chat/", json=payload, timeout=timeout)
        elapsed = time.perf_counter() - started
        status_code = response.status_code
        try:
            data = response.json()
        except ValueError:
            error = response.text[:300]
    except Exception as exc:  # noqa: BLE001 - report evidence should capture the exact failure.
        elapsed = time.perf_counter() - started
        error = f"{type(exc).__name__}: {exc}"

    answer = str(data.get("response") or data.get("answer") or "")
    trace_events = data.get("trace_events") or []
    sources = data.get("sources") or []
    route, intent = trace_summary(trace_events if isinstance(trace_events, list) else [])

    source_blob = " ".join(
        str(source.get("title") or "")
        + " "
        + str(source.get("file_name") or "")
        + " "
        + str(source.get("file_path") or "")
        + " "
        + str(source.get("source_url") or "")
        + " "
        + str(source.get("snippet") or "")
        for source in sources
        if isinstance(source, dict)
    )
    searchable_output = f"{answer} {source_blob}".lower()
    expected_keywords = split_keywords(case.get("expected_keywords", ""))
    matched_keywords = [keyword for keyword in expected_keywords if keyword.lower() in searchable_output]

    answered = status_code == 200 and bool(answer.strip())
    return {
        "id": case["id"],
        "category": case["category"],
        "query": case["query"],
        "expected_route": case.get("expected_route", ""),
        "actual_route": route,
        "actual_intent": intent,
        "status_code": status_code,
        "result": "PASS" if answered else "FAIL",
        "latency_seconds": round(elapsed, 3),
        "response_chars": len(answer),
        "source_count": len(sources),
        "matched_keywords": " | ".join(matched_keywords),
        "matched_keyword_count": len(matched_keywords),
        "expected_keyword_count": len(expected_keywords),
        "reference_source": case.get("reference_source", ""),
        "notes": case.get("notes", ""),
        "llm_provider": data.get("llm_provider") or provider,
        "llm_model": data.get("llm_model") or model,
        "error": error,
        "response_preview": " ".join(answer.split())[:240],
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, Any]], summary_rows: list[tuple[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for metric, value in summary_rows:
        summary_sheet.append([metric, value])

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
    summary_sheet.column_dimensions["A"].width = 36
    summary_sheet.column_dimensions["B"].width = 24

    result_sheet = workbook.create_sheet("API Results")
    result_sheet.append(list(rows[0].keys()))
    for row in rows:
        result_sheet.append(list(row.values()))

    for cell in result_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="CFE2F3")
    result_sheet.freeze_panes = "A2"
    for index, column_cells in enumerate(result_sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        result_sheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 10), 48)

    workbook.save(path)
    return True


def write_summary(path: Path, summary_rows: list[tuple[str, Any]], rows: list[dict[str, Any]]) -> None:
    passed = sum(1 for row in rows if row["result"] == "PASS")
    lines = [
        "# API 50-Question Test Evidence",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|---|---:|",
    ]
    for metric, value in summary_rows:
        lines.append(f"| {metric} | {value} |")

    lines.extend(
        [
            "",
            "## Notes",
            "",
            f"- All {passed} successful cases returned HTTP 200 with a non-empty API response.",
            "- Latency is measured client-side from the selected API transport.",
            "- Keyword matching is supporting evidence only; PASS/FAIL in this file means the API returned a usable response.",
            "- Open the CSV file directly in Excel, LibreOffice Calc, or Google Sheets for presentation.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate API evidence files for the 50-question test set.")
    parser.add_argument("--cases", default=str(DEFAULT_CASES))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--base-url", default="http://127.0.0.1:8001")
    parser.add_argument("--provider", default="openai")
    parser.add_argument("--model", default="gpt-5.4-mini")
    parser.add_argument("--timeout", type=float, default=120.0)
    parser.add_argument(
        "--transport",
        choices=("http", "testclient"),
        default="http",
        help="Use HTTP against a running backend or FastAPI TestClient with rate limiting disabled.",
    )
    args = parser.parse_args()

    cases = load_cases(Path(args.cases))
    generated_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    slug = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir)
    test_client = None
    if args.transport == "testclient":
        from fastapi.testclient import TestClient

        from src.backend.api.main import app

        app.state.limiter.enabled = False
        test_client = TestClient(app)

    rows = []
    for case in cases:
        row = run_case(
            case,
            base_url=args.base_url,
            provider=args.provider,
            model=args.model,
            timeout=args.timeout,
            transport=args.transport,
            test_client=test_client,
        )
        rows.append(row)
        print(
            f"[{row['result']}] {row['id']:>2} | "
            f"{row['latency_seconds']:.3f}s | sources={row['source_count']} | {row['query']}",
            flush=True,
        )

    latencies = [float(row["latency_seconds"]) for row in rows if row["result"] == "PASS"]
    passed = sum(1 for row in rows if row["result"] == "PASS")
    summary_rows: list[tuple[str, Any]] = [
        ("Generated at", generated_at),
        ("Base URL", args.base_url),
        ("Transport", args.transport),
        ("Total questions", len(rows)),
        ("Answered successfully", passed),
        ("Success rate", f"{passed / len(rows) * 100:.2f}%"),
        ("Average latency seconds", f"{statistics.mean(latencies):.2f}" if latencies else "0.00"),
        ("Median latency seconds", f"{statistics.median(latencies):.2f}" if latencies else "0.00"),
        ("Min latency seconds", f"{min(latencies):.2f}" if latencies else "0.00"),
        ("Max latency seconds", f"{max(latencies):.2f}" if latencies else "0.00"),
        ("Provider", args.provider),
        ("Model", args.model),
    ]

    csv_path = output_dir / f"api_50_question_results_{slug}.csv"
    xlsx_path = output_dir / f"api_50_question_results_{slug}.xlsx"
    md_path = output_dir / f"api_50_question_summary_{slug}.md"
    latest_csv = output_dir / "api_50_question_results_latest.csv"
    latest_xlsx = output_dir / "api_50_question_results_latest.xlsx"
    latest_md = output_dir / "api_50_question_summary_latest.md"

    write_csv(csv_path, rows)
    write_csv(latest_csv, rows)
    xlsx_written = write_xlsx(xlsx_path, rows, summary_rows)
    if xlsx_written:
        write_xlsx(latest_xlsx, rows, summary_rows)
    write_summary(md_path, summary_rows, rows)
    write_summary(latest_md, summary_rows, rows)

    print("\nGenerated files:")
    print(f"- {csv_path}")
    print(f"- {latest_csv}")
    if xlsx_written:
        print(f"- {xlsx_path}")
        print(f"- {latest_xlsx}")
    else:
        print("- XLSX skipped because openpyxl is not installed")
    print(f"- {md_path}")
    print(f"- {latest_md}")
    return 0 if passed == len(rows) else 1


if __name__ == "__main__":
    raise SystemExit(main())
